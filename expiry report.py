#split Expiry report by PMs and put in to tabs

import openpyxl

wb = openpyxl.load_workbook('Expiry.xlsx')
ws = wb.active
rngs = list(ws.values)
d = {}
for row in rngs[1:]:
    if row[8] in d.keys():
        d[row[8]] += [row]
    else:
        d[row[8]] = [row]

nwb = openpyxl.Workbook()
# key is PM , Position is Item
for pm, position in sorted(d.items()):
    nws=nwb.create_sheet(pm)  # create workbook
    nws.append(rngs[0])
    for expiry in position:
        nws.append(expiry)
nwb.remove(nwb.worksheets[0])
nwb.save('Expirydetails.xlsx')
